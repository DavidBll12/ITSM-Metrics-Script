## Program to create the base for the Weekly ITSM Problem Management Metric Report
## 'CW.xlsx' and 'RAW.xlsx' need to be in the working directory for program to work

import pandas as pd # Imported for filters mostly
import openpyxl # Imported for editing excel workbooks/worksheets
from datetime import datetime, timedelta # Imported for datetime objects used in data filter
from openpyxl.styles import Alignment # Imported to apply formatting to worksheets
from openpyxl.worksheet.table import Table, TableStyleInfo # Imported to apply formatting to worksheets
from openpyxl.utils.dataframe import dataframe_to_rows # Imported to convert Pandas dataframe for passing into worksheet

# Defines Workbooks 'CW.xlsx', and 'RAW.xslx', 'wb' == 'CW.xlsx', 'raw_wb' == 'RAW.xlsx'
# Defines Worksheets for "PRBs", "Raw Data" from "CW and "Page 1" from "RAW"
wb = openpyxl.load_workbook('CW.xlsx')
raw_wb = openpyxl.load_workbook('RAW.xlsx')
prb_sheet = wb["PRBs"]
raw_sheet = wb["Raw Data"]
sheet_from_raw = raw_wb["Page 1"]


# Defines the dates and path for the "closed_filter" function
# CHANGE THESE DATES BELOW TO REFLECT THIS WEEKS DATA, EX. SAT - FRI
# "end_date" SHOULD BE THE DAY AFTER THE DESIRED DATE, TO INCLUDE THE END DATE, EX. 28th -> 29th
start_date = '2023-04-22'
end_date = '2023-04-29'
file_path = 'finished_metrics.xlsx'

# Defines the report date for "report date" substitution:
# CHANGE THIS TO THE LAST DAY OF THE CURRENT WEEK OF DATA
report_date = "4/28/2023"

# Table number incrementation
table_num = 1 

# Problem Team concatetenation and Org mapping
# UPDATE TEAM NAMES AS NEEDED HERE, EX. 'name found in "Problem Team" ': ('team concatenation', 'team to map to')
team_map = {
        'Long Name': ('Concatenation', 'Department'),    
        'Network Operations': ('Network', 'Infra'),
    }

# Deletes the following sheets: Metrics - Weekly | Weekly Trend - Overall | Redacted | Redacted
# Redacted | Redacted | Redacted | Redacted

delete_sheets = ['Metrics - Weekly', 'Weekly Trend - Overall', 'Redacted', 
                    'Redacted', 'Redacted', 'Redacted', 'Redacted', 'Redacted']

# Defines the formulas for Column AF-AN, f string so that the formula matches the row it is in            
AF_formula = '=IF(AND(A{row}>"",Y{row}=0),DAYS360(X{row},AH{row}),"")'
AG_formula = '=IF(Y{row}<>0,DAYS360(X{row},Y{row}),"")'
AH_formula = '3/24/2023'
AI_formula = '=IF(AND(A{row}>"",Y{row}=0),IF(AF{row}<11,"( 00-10 )",IF(AF{row}<21,"( 11-20 )",IF(AF{row}<31,"( 21-30 )",IF(AF{row}<46,"( 31-45 )",IF(AF{row}<61,"( 46-60 )",IF(AF{row}<91,"( 61-90 )","( 90+ )")))))),"")'
AJ_formula = '=IF(C{row}>"",CONCATENATE(IF(C{row}="Proactive Problem","PP",IF(C{row}="Reactive Problem - EIM","EIM RP","RP"))," - ",AE{row}),"")'
AK_formula = '=IF(C{row}>"",CONCATENATE(IF(C{row}="Proactive Problem","PP","RP")," - ",AE{row}),"")'
AL_formula = '=IF(AJ{row}>"",CONCATENATE(AJ{row},"/",AD{row}),"")'
AM_formula = '=IF(AE{row}>"",CONCATENATE(AE{row},"/",AD{row},"/",AI{row}),"")'
AN_formula = '=IF(C{row}>"",CONCATENATE(IF(C{row}="Proactive Problem","PP",IF(C{row}="Reactive Problem - EIM","EIM RP","RP")),"/",AM{row}),"")'

############################################EMPTY CELLS#######################################################################
# Replaces empty cells with the string "TBD"

# Defines a function that replaces empty cells in a given range with the string "TBD"
def replace_empty_cells(dest_sheet, col_start, col_end):
    # Iterate over each row in the worksheet, starting from row 2
    for row in range(2, dest_sheet.max_row + 1):
        col1_cell = dest_sheet.cell(row=row, column=1)
        # Check if the value in the first column cell is not empty or NaN
        if col1_cell.value is not None or pd.notna(col1_cell.value) or (isinstance(col1_cell.value, float) and col1_cell.value == col1_cell.value):
            # Iterate over each column in the given range
            for col in range(col_start, col_end):
                # Get the value of the current cell
                cell = dest_sheet.cell(row=row, column=col)
                value = cell.value
                
                # Check if the value is empty or NaN or a float NaN value
                if value is None or (isinstance(value, float) and value != value) or pd.isna(value):
                    # Replace the empty cell with the string "TBD"
                    cell.value = 'TBD'
                
    # Saves the modiefied worksheet                
    wb.save('finished_metrics.xlsx')


############################################sheet_edit#######################################################################
def sheet_edit(dest_sheet):
    # Loop through each row in "State" column and update corresponding values in other columns
    for row in range(2, dest_sheet.max_row + 1):
        # Get the value in "State" column
        state_val = dest_sheet.cell(row=row, column=6).value

        # Update values in columns AB, AA, AD, and AE based on the value in "State" column
        if state_val == 'Resolved/Closed':
            dest_sheet.cell(row=row, column=28).value = 'Resolved'
        elif state_val == 'Cancelled/Closed':
            dest_sheet.cell(row=row, column=27).value = 'N/A'
            dest_sheet.cell(row=row, column=28).value = 'Cancelled'
        elif state_val == 'Known Error/Closed':
            dest_sheet.cell(row=row, column=27).value = 'N/A'
            dest_sheet.cell(row=row, column=28).value = 'KE/WO'

        # Checks to see if there is a value in col1 and if col31 is "TBD" or empty it replaces it with "Other"    
        if dest_sheet.cell(row=row, column=1).value != None and (dest_sheet.cell(row=row, column=31).value == "TBD" or dest_sheet.cell(row=row, column=31).value is None):
            dest_sheet.cell(row=row, column=31).value = 'Other'       

        # Iterate "Problem Team" and replace with "team_map" values
        long_team_name = dest_sheet.cell(row=row, column=30).value
        if long_team_name in team_map:
            concat_team_name, org_name = team_map[long_team_name]
            dest_sheet.cell(row=row, column=30).value = concat_team_name
            dest_sheet.cell(row=row, column=31).value = org_name

    # Checks through values in column "AA", and assigns them to a "float" if possible. If the values are not none. 
    for cell in dest_sheet['AA']:
        if cell.value != None:
            try:
                cell.value = float(cell.value)
            except (TypeError, ValueError):
                cell.value = cell.value

    # Loop through each row in the worksheet
    for row in range(2, dest_sheet.max_row + 1):
        # Check if column A is not None type
        if dest_sheet[f'A{row}'].value is not None:
            # Apply the formulas to the respective columns
            dest_sheet[f'AF{row}'] = AF_formula.format(row=row)
            dest_sheet[f'AG{row}'] = AG_formula.format(row=row)
            dest_sheet[f'AH{row}'] = AH_formula.format(row=row)
            dest_sheet[f'AI{row}'] = AI_formula.format(row=row)
            dest_sheet[f'AJ{row}'] = AJ_formula.format(row=row)
            dest_sheet[f'AK{row}'] = AK_formula.format(row=row)
            dest_sheet[f'AL{row}'] = AL_formula.format(row=row)
            dest_sheet[f'AM{row}'] = AM_formula.format(row=row)
            dest_sheet[f'AN{row}'] = AN_formula.format(row=row)

    for row in range(2, dest_sheet.max_row + 1):
        # Check if column A is not None type
        if dest_sheet[f'A{row}'].value is not None or (isinstance(value, float) and value != value) or pd.isna(value):            
            dest_sheet.cell(row=row, column=34).value = report_date
    
            
            
# Loops through the sheets in "delete_sheets" and deletes them
for sheet in delete_sheets:
    del wb[sheet]

# Loops through all cells in "PRBs" and deletes them
for row_num in range(2, prb_sheet.max_row + 1):
    for col_num in range(1, prb_sheet.max_column + 1):
        prb_sheet.cell(row = row_num, column = col_num).value = None

# Loops through all cells in "Raw Data" and deletes them
for row_num in range(2, raw_sheet.max_row + 1):
    for col_num in range(1, raw_sheet.max_column + 1):
        raw_sheet.cell(row = row_num, column = col_num).value = None

# Copy everything from "RAW.xlsx" "Page 1" except the header/row1 without formatting
# Append that coppied data into the "copied_rows" list
copied_rows = []
for row in sheet_from_raw.iter_rows(min_row=2, values_only=True):
    copied_rows.append(row)

# Paste the copied rows into cell A2 of "raw_sheet"
for row_idx, row in enumerate(copied_rows, start=2):
    for col_idx, value in enumerate(row, start=1):
        raw_sheet.cell(row=row_idx, column=col_idx).value = value

wb.save("finished_metrics.xlsx")


#################################################CLOSED FILTER#################################################################
def closed_filter(file_path, start_date, end_date, dest_sheet):
    
    # Loads data from Excel file
    df = pd.read_excel(file_path, sheet_name='Raw Data')
    
    # Selects the first 31 columns of the DataFrame
    df = df.iloc[:, :31]
    
    # Convert 'Closed' column to datetime format
    df['Closed'] = pd.to_datetime(df['Closed'], errors='coerce')

    # Filter data for dates between start_date and end_date
    mask = ((df['Closed'] >= start_date) & (df['Closed'] <= end_date)) | (df['Closed'].isna())
    
    filtered_data = df.loc[mask]
    
    # Convert filtered data to rows
    rows = dataframe_to_rows(filtered_data, index=False, header=False)
    
    # Append rows to destination sheet
    for row_idx, row in enumerate(rows, 1):
        for col_idx, value in enumerate(row, 1):
            dest_sheet.cell(row=row_idx+1, column=col_idx, value=value)
            
    wb.save("finished_metrics.xlsx")

closed_filter(file_path, start_date, end_date, prb_sheet)   

# Reasigned values since they werent working
wb = openpyxl.load_workbook('finished_metrics.xlsx')
prb_sheet = wb["PRBs"]
raw_sheet = wb["Raw Data"]
        
# Call the function with the correct arguments
sheet_edit(prb_sheet)

# Call the function with the correct arguments
replace_empty_cells(prb_sheet, 27, 31)   

# Defines sheets and workbooks for "sheet_created_closed" and "sheet_open" functions
created_sheet = wb.create_sheet("Created",0)
closed_sheet = wb.create_sheet("Closed",1)
open_sheet = wb.create_sheet("Open",2)     

# Save so that PRB is the same as Created,Closed,Open
wb.save("finished_metrics.xlsx")

#########################################CREATED AND CLOSED SHEETS#############################################################
def sheet_created_closed(file_path, start_date, end_date, dest_sheet, column_name):  
    
    # Load data from Excel file
    df = pd.read_excel(file_path, sheet_name='PRBs', keep_default_na=False)

    # Convert 'Created' column to datetime format
    df[column_name] = pd.to_datetime(df[column_name], errors='coerce')

    # Filter data for dates between start_date and end_date
    mask = (df[column_name] >= pd.Timestamp(start_date)) & (df[column_name] <= pd.Timestamp(end_date))
    filtered_data = df.loc[mask]
    
    # Convert filtered data to rows
    rows = dataframe_to_rows(filtered_data, index=False, header=True)
    
    # Append rows to destination sheet
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            dest_sheet.cell(row=row_idx, column=col_idx).value = value
            
    # Determine the range of populated cells
    max_row = dest_sheet.max_row
    max_column = dest_sheet.max_column
    min_row = dest_sheet.min_row
    min_column = dest_sheet.min_column
   
    # Set cell alignment, wrapping and row height
    for row in dest_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_column, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        dest_sheet.row_dimensions[cell.row].height = 55
        
    # Set column width for all columns
    for col in dest_sheet.columns:
        dest_sheet.column_dimensions[col[0].column_letter].width = 15
    
    # Set the first row height to "30"
    dest_sheet.row_dimensions[1].height = 30   
    
    # Adjust specific column widths for best viewing
    for col in ["A", "B", "C", "F"]:
        dest_sheet.column_dimensions[col].width = 13
    for col in ["E", "J", "K", "T"]:
        dest_sheet.column_dimensions[col].width = 45
    for col in ["AJ", "AK", "AL", "AM", "AN"]:
        dest_sheet.column_dimensions[col].width = 30
        
    dest_sheet.column_dimensions["D"].width = 8
    dest_sheet.column_dimensions["H"].width = 17

    # Create a table
    # Update the "ref" parameter to include the determined range of cells
    table_ref = f"{dest_sheet.cell(row=min_row, column=min_column).coordinate}:{dest_sheet.cell(row=max_row, column=max_column).coordinate}"
  
    # Incremement table numbers FUTURE: implement in a create table function
    global table_num
    table_name = f"Table{table_num}"
    table = Table(displayName=table_name, ref=table_ref)
    table_num += 1  # increment table number for the next iteration
    
    # Apply the "Blue, Table Style Medium 2" format
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    dest_sheet.add_table(table) 
    
    # Set sheet zoom to 90
    dest_sheet.sheet_view.zoomScale = 90
    
    # Freeze columns A-F
    dest_sheet.freeze_panes = 'G1'
    
    # Loop through each row in the worksheet
    for row in range(2, dest_sheet.max_row + 1):
        dest_sheet.cell(row=row, column=32).value = AF_formula.format(row=row)
        dest_sheet.cell(row=row, column=33).value = AG_formula.format(row=row)
        dest_sheet.cell(row=row, column=34).value = AH_formula
        dest_sheet.cell(row=row, column=35).value = AI_formula.format(row=row)
        dest_sheet.cell(row=row, column=36).value = AJ_formula.format(row=row)
        dest_sheet.cell(row=row, column=37).value = AK_formula.format(row=row)
        dest_sheet.cell(row=row, column=38).value = AL_formula.format(row=row)
        dest_sheet.cell(row=row, column=39).value = AM_formula.format(row=row)
        dest_sheet.cell(row=row, column=40).value = AN_formula.format(row=row)

    for row in range(2, dest_sheet.max_row + 1):
        # Check if column A is not None type
        if dest_sheet[f'A{row}'].value is not None:            
            dest_sheet.cell(row=row, column=34).value = report_date   

    # Save the changes to the Excel file
    wb.save("finished_metrics.xlsx")
    
############################################OPEN SHEETS#######################################################################    
def sheet_open(file_path,dest_sheet, column_name):
    
    # Load data from Excel file
    df = pd.read_excel(file_path, sheet_name='PRBs', keep_default_na=False)

    # Filter data for dates between start_date and end_date
    mask = ~(df[column_name].str.contains('Resolved/Closed') | df[column_name].str.contains('Known Error/Closed') | df[column_name].str.contains('Cancelled/Closed'))
    filtered_data = df.loc[mask]
    
    # Convert filtered data to rows
    rows = dataframe_to_rows(filtered_data, index=False, header=True)
    
    # Append rows to destination sheet
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            dest_sheet.cell(row=row_idx, column=col_idx).value = value
            
    # Determine the range of populated cells
    max_row = dest_sheet.max_row
    max_column = dest_sheet.max_column
    min_row = dest_sheet.min_row
    min_column = dest_sheet.min_column
        
    # Set cell alignment, wrapping and row height
    for row in dest_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_column, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        dest_sheet.row_dimensions[cell.row].height = 55
        
    # Set column width for all columns
    for col in dest_sheet.columns:
        dest_sheet.column_dimensions[col[0].column_letter].width = 15
    
    # Set the first row height to "30"
    dest_sheet.row_dimensions[1].height = 30   
    
    # Adjust specific column widths for best viewing
    for col in ["A", "B", "C", "F"]:
        dest_sheet.column_dimensions[col].width = 13
    for col in ["E", "J", "K", "T"]:
        dest_sheet.column_dimensions[col].width = 45
    for col in ["AJ", "AK", "AL", "AM", "AN"]:
        dest_sheet.column_dimensions[col].width = 30
        
    dest_sheet.column_dimensions["D"].width = 8
    dest_sheet.column_dimensions["H"].width = 17
       
    # Create a table
    # Update the "ref" parameter to include the determined range of cells
    table_ref = f"{dest_sheet.cell(row=min_row, column=min_column).coordinate}:{dest_sheet.cell(row=max_row, column=max_column).coordinate}"
    
    # Defines 'table_num' as a global variable, used because the variable is defined outside of this funciton
    global table_num
    # Defined the table name because functions in the Workbook rely on Table numbers being constant
    table_name = f"Table{table_num}"
    table = Table(displayName=table_name, ref=table_ref)
    table_num += 1  # increment table number for the next iteration
    
    # Apply the "Blue, Table Style Medium 2" format
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    dest_sheet.add_table(table)  
    
    # Set sheet zoom to 90
    dest_sheet.sheet_view.zoomScale = 90
    
    # Freeze columns A-F
    dest_sheet.freeze_panes = 'G1'

    # Loop through each row in the worksheet
    for row in range(2, dest_sheet.max_row + 1):
        dest_sheet.cell(row=row, column=32).value = AF_formula.format(row=row)
        dest_sheet.cell(row=row, column=33).value = AG_formula.format(row=row)
        dest_sheet.cell(row=row, column=34).value = AH_formula
        dest_sheet.cell(row=row, column=35).value = AI_formula.format(row=row)
        dest_sheet.cell(row=row, column=36).value = AJ_formula.format(row=row)
        dest_sheet.cell(row=row, column=37).value = AK_formula.format(row=row)
        dest_sheet.cell(row=row, column=38).value = AL_formula.format(row=row)
        dest_sheet.cell(row=row, column=39).value = AM_formula.format(row=row)
        dest_sheet.cell(row=row, column=40).value = AN_formula.format(row=row)

    for row in range(2, dest_sheet.max_row + 1):
        # Check if column A is not None type
        if dest_sheet[f'A{row}'].value is not None:            
            dest_sheet.cell(row=row, column=34).value = report_date     
        
# Call the function with the correct arguments
sheet_created_closed(file_path, start_date, end_date, created_sheet, "Created")
sheet_created_closed(file_path, start_date, end_date, closed_sheet, "Closed")
sheet_open(file_path,open_sheet, "State")


# Defines sheets and workbooks for "sheet_created_closed" and "sheet_open" functions
ppm_created_sheet = wb.create_sheet("PPM Created",5)
ppm_closed_sheet = wb.create_sheet("PPM Closed",6)
     
# Save the changes to the Excel file
wb.save("finished_metrics.xlsx")      

def ppm_created(file_path, dest_sheet, column_name, ins_col_num, ins_col_name, formula_letter):  
    
    # Load data from Excel file
    df = pd.read_excel(file_path, sheet_name='Raw Data', keep_default_na=False)
    
    # Convert 'Created' column to datetime format
    df[column_name] = pd.to_datetime(df[column_name], errors='coerce')

    # Filter data for dates between start_date and end_date, and for specific organization and problem type
    mask = (df[column_name].dt.year == 2023) & (df["Organization"].isin(["Infra", "AMS"])) & (df["Problem Type"] == "Proactive Problem") & (df['State'] != "Cancelled/Closed")
    filtered_data = df.loc[mask]
    filtered_data = filtered_data.loc[:, ~filtered_data.columns.str.startswith('Unnamed')]
    
    # Convert filtered data to rows
    rows = dataframe_to_rows(filtered_data, index=False, header=True)
    
    # Append rows to destination sheet
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            dest_sheet.cell(row=row_idx, column=col_idx).value = value
            
    for row in range(2, dest_sheet.max_row + 1):
        
        # Get the value in "State" column
        state_val = dest_sheet.cell(row=row, column=6).value

        # Update values in columns AB, AA, AD, and AE based on the value in "State" column
        if state_val == 'Known Error/Closed':
            dest_sheet.cell(row=row, column=27).value = 'N/A'
            dest_sheet.cell(row=row, column=28).value = 'KE/WO'
        if state_val == 'Resolved/Closed':
            dest_sheet.cell(row=row, column=28).value = 'Resolved'
        
        # Iterate "Problem Team" and replace with "team_map" values
        long_team_name = dest_sheet.cell(row=row, column=30).value
        if long_team_name in team_map:
            concat_team_name, org_name = team_map[long_team_name]
            dest_sheet.cell(row=row, column=30).value = concat_team_name
            dest_sheet.cell(row=row, column=31).value = org_name
            
    # insert a new column at column defined by ins_col_num
    dest_sheet.insert_cols(ins_col_num)

    # set the header for the new column
    dest_sheet.cell(row=1, column= ins_col_num, value= ins_col_name)

    # set the formula for the new column
    for row in range(2, dest_sheet.max_row + 1):
        dest_sheet.cell(row=row, column=ins_col_num).value = '=TEXT(' + formula_letter + '{},"mmmm")'.format(row)
        
    # Determine the range of populated cells
    max_row = dest_sheet.max_row
    max_column = dest_sheet.max_column
    min_row = dest_sheet.min_row
    min_column = dest_sheet.min_column
        
    # Set cell alignment, wrapping and row height
    for row in dest_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_column, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        dest_sheet.row_dimensions[cell.row].height = 55
        
    # Set column width for all columns
    for col in dest_sheet.columns:
        dest_sheet.column_dimensions[col[0].column_letter].width = 15
    
    # Set the first row height to "30"
    dest_sheet.row_dimensions[1].height = 30   
    
    # Adjust specific column widths for best viewing
    for col in ["A", "B", "C", "F"]:
        dest_sheet.column_dimensions[col].width = 13
    for col in ["E", "J", "K", "T"]:
        dest_sheet.column_dimensions[col].width = 45
        
    dest_sheet.column_dimensions["D"].width = 8
    dest_sheet.column_dimensions["H"].width = 17
       
    # Create a table
    # Update the "ref" parameter to include the determined range of cells
    table_ref = f"{dest_sheet.cell(row=min_row, column=min_column).coordinate}:{dest_sheet.cell(row=max_row, column=max_column).coordinate}"
    # table = Table(displayName=f"{dest_sheet.title}_Table", ref=table_ref)
    
    global table_num
    table_name = f"Table{table_num}"
    table = Table(displayName=table_name, ref=table_ref)
    table_num += 1  # increment table number for the next iteration
    
    # Apply the "Blue, Table Style Medium 2" format
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    dest_sheet.add_table(table)  
    
    # Set sheet zoom to 90
    dest_sheet.sheet_view.zoomScale = 90
    
    # Freeze columns A-F
    dest_sheet.freeze_panes = 'G1'    

    replace_empty_cells(dest_sheet, 28, 31) 

    wb.save("finished_metrics.xlsx")  
    
# Call the function with the correct arguments
ppm_created(file_path, ppm_created_sheet, "Created", 25, "Created Month", "X")
ppm_created(file_path, ppm_closed_sheet, "Closed", 26, "Closed Month", "Y")

# #Save the destination workbook
wb.save('finished_metrics.xlsx')

